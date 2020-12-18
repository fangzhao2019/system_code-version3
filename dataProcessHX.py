# coding:utf-8
import pymysql
import numpy as np
from pymongo import MongoClient
from openpyxl import load_workbook
import jieba
import jieba.posseg as pseg
jieba.load_userdict('model/userdict.txt')
import re

#载入产品特征集合
def loadFeatureSet(filename):
    featureSet={}
    f=open(filename,encoding='utf-8')
    for row in f.readlines()[1:]:
        data=row.replace('\n','').split('\t')
        featureSet[data[0]]=data[1]
    return featureSet

#将长句切分为短句
def clauseSegmentation(comment):
    reg1=re.compile(u'。+')
    reg2=re.compile(u'[？?]+')
    reg3=re.compile(u'[！!]+')
    reg4=re.compile(u'[;；]+')
    reg5=re.compile(u'\d[、,，：:]+')
    reg6=re.compile(u'…+')
    reg7=re.compile(u'（\d）')
    #reg9=re.compile(u'[，,]+')
    reg10=re.compile(u'\t+')
    comment=reg1.sub(u'。【Instead】',comment)
    comment=reg2.sub(u'？【Instead】',comment)
    comment=reg3.sub(u'！【Instead】',comment)
    comment=reg4.sub(u'；【Instead】',comment)
    comment=reg5.sub(u'【Instead】',comment)
    comment=reg6.sub(u'……【Instead】',comment)
    comment=reg7.sub(u'【Instead】',comment)
    #comment=reg9.sub('，【Instead】',comment)
    comment=reg10.sub('【Instead】',comment)
    clauseSet=re.split(u'【Instead】',comment)
    clauseList=[clause for clause in clauseSet if len(clause)>5]
    return clauseList

def load_carNameSet():
    wb=load_workbook('model/车名整理.xlsx')
    ws=wb.active
    extra_carNameSet=[]
    extra_carNameDic={}
    inter_carNameDic={}
    
    for i in range(2,ws.max_row+1):
        car=ws.cell(row=i,column=1).value.upper()
        brand=ws.cell(row=i,column=2).value
        inter_names=ws.cell(row=i,column=3).value.split(',')
        extra_names=ws.cell(row=i,column=4).value.split(',')
        inter_carNameDic[car]=list(set(inter_names)|set([brand]))
        for nn in extra_names:
            extra_carNameSet.append(nn)
            extra_carNameDic[nn]=car
    return extra_carNameSet,extra_carNameDic,inter_carNameDic

def carIdentify(inter_carNameSet,commentVec,tag):                  
    for i in range(len(commentVec)):
        if commentVec[i] in inter_carNameSet:
            tag[i]='INCAR'
    return tag

def compareCarIdentify(extra_carNameSet,commentVec,tag):
    for i in range(len(commentVec)):
        if commentVec[i] in extra_carNameSet:
            if tag[i]==0:
                tag[i]='EXCAR'
    return tag

def opinionIdentify(featureSet,senWords,notWords,commentVec,tag):
    for i in range(len(commentVec)):
        if commentVec[i] in featureSet:
            if tag[i]==0:
                tag[i]='FEATURE'
        if commentVec[i] in senWords:
            if tag[i]==0:
                tag[i]='SENWORD'
        if commentVec[i] in notWords:
            if tag[i]==0:
                tag[i]='NOTWORD'
    return tag        

#找到某个特征词对应的所有位置
def myfind(x,y):
    return [ a for a in range(len(y)) if y[a] == x]

def resultProcess(car,extra_carNameDic,commentVec,commentPOS,tag,featureOpinion):

    resultSet=[]
    fea_indexs=myfind('FEATURE',tag)
    inc_indexs=myfind('INCAR',tag)
    exc_indexs=myfind('EXCAR',tag)      
            
    if len(exc_indexs)>0:
        for index in fea_indexs:
            #特征描述了前面最近的一个车型
            
            for i in range(index,index-20,-1):
                if i<0:break
                if commentVec[i]=='。':break
                #提及当前车型
                if tag[i]=='INCAR':
                    break
                if tag[i]=='EXCAR':
                    count=0
                    opinion=commentVec[index]
                    for j in range(index-1,max(i,index-4),-1):
                        if commentPOS[j]=='x':break
                        if j in fea_indexs:break
                        if  tag[j] in ['SENWORD','NOTWORD']:
                            opinion=opinion+commentVec[j]
                            count+=1
                    if count==0:
                        for j in range(index+1,index+6):
                            if j>=len(tag):break
                            if commentPOS[j]=='x':break
                            if j in fea_indexs:break
                            if  tag[j] in ['SENWORD','NOTWORD']:
                                opinion=opinion+commentVec[j]
                                count+=1
                    if count==0:
                        resultSet.append([extra_carNameDic[commentVec[i]],commentVec[index],''])
                    if count>0:
                        if opinion in featureOpinion:
                            resultSet.append([extra_carNameDic[commentVec[i]],commentVec[index],opinion])
                        else:
                            resultSet.append([extra_carNameDic[commentVec[i]],commentVec[index],''])
                    break
    #去重
    new_resultSet=[]
    for r in resultSet:
        if not r in new_resultSet:
            new_resultSet.append(r)
    return new_resultSet

def splitDataSet(dataSet, axis, value):#按照给定特征划分数据集
    retDataSet = []#用来存储划分结果的列表
    for featVec in dataSet:#对于每行数据
        if featVec[axis] == value:#如果选择的特征等于给定值
            reducedFeatVec = featVec[:axis]
            reducedFeatVec.extend(featVec[axis+1:])
            retDataSet.append(reducedFeatVec)#将这行数据存入划分结果的列表中（不包含给定特征）
    return retDataSet#返回划分结果数据集

def formatDataSet(opinionSet):
    formatOpinionSet={}
    feaUniqueVals=set([exa[0] for exa in opinionSet])
    for value1 in feaUniqueVals:
        subDataSet=splitDataSet(opinionSet,0,value1)
        senUniqueVals=set([exa[0] for exa in subDataSet])
        subDic={}
        for value2 in senUniqueVals:
            commentSet=[c[0] for c in splitDataSet(subDataSet,0,value2)]
            subDic[value2]=commentSet
        keys=subDic.keys()
        if '' in keys:
            if len(keys)>1:
                subDic.pop('')
            if len(keys)==1:
                subDic={}
        formatOpinionSet[value1]=subDic
    return formatOpinionSet

    
############################数据处理############################

client=MongoClient('192.168.1.94',30000)
collection1=client.car_autohome.comment_user
collection2=client.car_dev.autohome_comment_user

featureSet=loadFeatureSet('model/featureSet.txt')
purchaceFactors=loadFeatureSet('model/purchace_factor.txt')
exteriorImage=loadFeatureSet('model/exteriorImage.txt')
interiorImage=loadFeatureSet('model/interiorImage.txt')

filterFeature=loadFeatureSet('model/filterFeature.txt')
featureOpinion=list(loadFeatureSet('model/featureOpinionSet.txt').keys())

featureSet1=list(featureSet.keys())
senWords=[w.replace('\n','') for w in open('model/senWords.txt',encoding='utf-8').readlines()[1:]]
degreeWords=[w.replace('\n','') for w in open('model/degreeWords.txt',encoding='utf-8').readlines()[1:]]
notWords=[w.replace('\n','') for w in open('model/notWords.txt',encoding='utf-8').readlines()[1:]]
extra_carNameSet,extra_carNameDic,inter_carNameDic=load_carNameSet()

print('一共包含待处理口碑数据%s条'%collection1.find().count())
print('一共包含产品特征%d个'%len(featureSet))
print('一共包含情感词%d个'%len(senWords))

i=0
for row in collection1.find({}):
    i+=1
    if i%10==0:print(i)
    if not 'data' in row.keys():continue
    
    new_mongoData={}
    #id
    new_mongoData['_id']=row['_id']
    
    #购买车型
    if 'modelName' in row.keys():
        carName=row['modelName']
        new_mongoData['carName']=carName
    else:
        continue

    if not carName in inter_carNameDic.keys():
        carName=carName.replace('-','')
        if not carName in inter_carNameDic.keys():
            continue
    
    data=row['data']

    #购买车款
    if '购买车型' in data.keys():
        modelName=data['购买车型']
        new_mongoData['modelName']=modelName
    elif 'editionName' in row.keys():
        modelName=row['editionName']
        new_mongoData['modelName']=modelName
    else:
        continue

    #排量
    new_mongoData['排量']=modelName.split(' ')[1]
    
    #购买地点
    if '购买地点' in data.keys():
        new_mongoData['购买地点']=data['购买地点']
    else:
        new_mongoData['购买地点']=''
        
    #购买时间
    if '购买时间' in data.keys():
        new_mongoData['购买时间']=data['购买时间']
    else:
        new_mongoData['购买时间']=''

    #购车目的
    if 'purchase_purpose' in row.keys():
        new_mongoData['购车目的']=row['purchase_purpose'].split('/')
    else:
        new_mongoData['购车目的']=[]
    
    
    if not '评价' in data.keys():continue
    if not '评分' in data.keys():continue
    comment=data['评价']
    
    #购买因素
    comment1=''
    for key in list(comment.keys()):
        if '为什么' in key:
            comment1=comment[key]
    if len(comment1)>0:
        comment1Vec=list(jieba.cut(comment1))
        factors=[]
        for  key in list(purchaceFactors.keys()):
            if key in comment1Vec:
                factor=purchaceFactors[key]
                if not factor in factors:
                    factors.append(factor)
    else:
        factors=[]
    new_mongoData['购买因素']=factors

    #外观形象
    comment2=[]
    for key in list(comment.keys()):
        if key=='外观':comment2.append(comment[key])
        if '最满意' in key:
            for clause in clauseSegmentation(comment[key]):
                if '外观' in clause:
                    comment2.append(clause)
    comment2='。'.join(comment2)
    if len(comment2)>0:
        comment2Vec=list(jieba.cut(comment2))
        ex_images=[]
        for key in list(exteriorImage.keys()):
            if key in comment2Vec:
                ex_image=exteriorImage[key]
                if not ex_image in ex_images:
                    ex_images.append(ex_image)
    else:
        ex_images=[]
    new_mongoData['外观形象']=ex_images

    #内饰形象
    comment3=[]
    for key in list(comment.keys()):
        if key=='内饰':comment3.append(comment[key])
        if '最满意' in key:
            for clause in clauseSegmentation(comment[key]):
                if '内饰' in clause:
                    comment3.append(clause)
    comment3='。'.join(comment3)
    if len(comment3)>0:
        comment3Vec=list(jieba.cut(comment3))
        in_images=[]
        for key in list(interiorImage.keys()):
            if key in comment3Vec:
                in_image=interiorImage[key]
                if not in_image in in_images:
                    in_images.append(in_image)
    else:
        in_images=[]
    new_mongoData['内饰形象']=in_images

    #卡诺模型
    comment4=''
    comment5=''
    for key in list(comment.keys()):
        if '最满意' in key:
            comment4=comment[key]
            break
    for key in list(comment.keys()):
        if '最不满意' in key:
            comment5=comment[key]
            break
    comment4Vec=list(jieba.cut(comment4))
    comment5Vec=list(jieba.cut(comment5))
    kano_single={}
    positive=[]
    negative=[]
    for fea in featureSet1:
        if fea in comment4Vec:
            positive.append(fea)
        if fea in comment5Vec:
            negative.append(fea)
    kano_single['正面情感']=list(set(positive))
    kano_single['负面情感']=list(set(negative))
    try:
        score=[int(w) for w in list(data['评分'].values())]
        kano_single['总分']=sum(score)/len(score)
    except:
        print('评分数据错误')
        continue
    new_mongoData['卡诺模型']=kano_single

    #竞争分析
    if len(comment1)>0:
        carName=carName.upper()
        commentVec=[w.word for w in pseg.cut(comment1)]
        commentPOS=[w.flag for w in pseg.cut(comment1)]
        tag=[0]*len(commentVec)
        tag=carIdentify(inter_carNameDic[carName],commentVec,tag)
        tag=compareCarIdentify(extra_carNameSet,commentVec,tag)
        tag=opinionIdentify(featureSet1,senWords,notWords,commentVec,tag)
        resultSet=resultProcess(carName,extra_carNameDic,commentVec,commentPOS,tag,featureOpinion)
        formatSet=formatDataSet(resultSet)
    else:
        formatSet={}
    new_mongoData['竞争分析']=formatSet
        
    try:
        collection2.insert(new_mongoData)
    except:
        continue
